"""
FACE ATTENDANCE SYSTEM WITH FIREBASE - MERGED FINAL
Gabungan fitur simple version + Firebase fixes:
1. Register face: Haarcascade + SPACE/Auto mode (dari simple)
2. Register unknown: name entry inline + tombol Delete Photo (dari simple)  
3. Attendance: ESC to stop, no time limit, unknown auto-save with dedup (dari simple)
4. Base64 encoding + user_ key prefix (Firebase fixes)
"""

import cv2
import numpy as np
import os
import json
import base64
from datetime import datetime
import tkinter as tk
from tkinter import messagebox, simpledialog
from PIL import Image, ImageTk
import firebase_admin
from firebase_admin import credentials, db
import time


class FirebaseManager:
    def __init__(self):
        self.initialized = False
        self.db_ref = None

    def initialize(self):
        try:
            config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'firebase-config.json')

            if not os.path.exists(config_path):
                self._create_template_config(config_path)
                messagebox.showwarning("Firebase Setup Required",
                    "firebase-config.json telah dibuat!\n\n"
                    "1. Buka Firebase Console\n"
                    "2. Project Settings > Service Accounts\n"
                    "3. Generate New Private Key\n"
                    "4. Replace isi firebase-config.json\n"
                    "5. Restart aplikasi.")
                return False

            with open(config_path, 'r') as f:
                config = json.load(f)

            if config.get('type') == 'REPLACE_WITH_YOUR_SERVICE_ACCOUNT_KEY':
                messagebox.showwarning("Firebase Belum Dikonfigurasi",
                    "Isi dulu firebase-config.json dengan kredensial Firebase kamu!")
                return False

            if not firebase_admin._apps:
                cred = credentials.Certificate(config_path)
                firebase_admin.initialize_app(cred, {
                    'databaseURL': config.get('databaseURL', '')
                })

            self.db_ref = db.reference()
            self.initialized = True
            print("✅ Firebase connected!")
            return True

        except Exception as e:
            print(f"❌ Firebase init failed: {e}")
            messagebox.showerror("Firebase Error", f"Gagal connect:\n{str(e)}")
            return False

    def _create_template_config(self, path):
        template = {
            "type": "REPLACE_WITH_YOUR_SERVICE_ACCOUNT_KEY",
            "project_id": "your-project-id",
            "private_key_id": "your-private-key-id",
            "private_key": "-----BEGIN PRIVATE KEY-----\\nYOUR_KEY\\n-----END PRIVATE KEY-----\\n",
            "client_email": "your-service-account@your-project.iam.gserviceaccount.com",
            "client_id": "your-client-id",
            "auth_uri": "https://accounts.google.com/o/oauth2/auth",
            "token_uri": "https://oauth2.googleapis.com/token",
            "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
            "client_x509_cert_url": "https://www.googleapis.com/robot/v1/metadata/x509/your-service-account",
            "databaseURL": "https://your-project-id.firebaseio.com"
        }
        with open(path, 'w') as f:
            json.dump(template, f, indent=2)

    def _make_key(self, user_id):
        """FIX: prefix user_ supaya Firebase tidak konvert dict -> list"""
        return f"user_{user_id}"

    def _encode_face(self, face_encoding):
        """FIX: compress numpy array -> base64 string (bukan 40.000 node individual)"""
        if isinstance(face_encoding, np.ndarray):
            face_bytes = face_encoding.tobytes()
            face_b64 = base64.b64encode(face_bytes).decode('utf-8')
            return {
                'encoding': face_b64,
                'shape': list(face_encoding.shape),
                'dtype': str(face_encoding.dtype)
            }
        return face_encoding

    def _decode_face(self, face_data):
        """Decode base64 -> numpy, support format lama juga"""
        if face_data is None:
            return None
        try:
            if isinstance(face_data, dict) and 'encoding' in face_data:
                face_bytes = base64.b64decode(face_data['encoding'])
                shape = tuple(face_data['shape'])
                dtype = np.dtype(face_data['dtype'])
                return np.frombuffer(face_bytes, dtype=dtype).reshape(shape)
            elif isinstance(face_data, list):
                enc = np.array(face_data, dtype=np.uint8)
                if len(enc) == 2500:
                    return cv2.resize(enc.reshape(50, 50), (200, 200))
                elif len(enc) == 40000:
                    return enc.reshape(200, 200)
                else:
                    print(f"⚠️  Unknown encoding size: {len(enc)}")
                    return None
        except Exception as e:
            print(f"⚠️  Decode error: {e}")
            return None

    def save_user(self, user_id, name, face_encoding):
        if not self.initialized:
            return False
        try:
            key = self._make_key(user_id)
            face_data = self._encode_face(face_encoding)
            self.db_ref.child('users').child(key).set({
                'name': name,
                'id': user_id,
                'face_encoding': face_data,
                'created_at': datetime.now().isoformat(),
                'updated_at': datetime.now().isoformat()
            })
            return True
        except Exception as e:
            print(f"❌ Save user failed: {e}")
            return False

    def get_user(self, user_id):
        if not self.initialized:
            return None
        try:
            return self.db_ref.child('users').child(self._make_key(user_id)).get()
        except Exception as e:
            print(f"❌ Get user failed: {e}")
            return None

    def get_all_users(self):
        if not self.initialized:
            return {}
        try:
            users = self.db_ref.child('users').get()
            return users if users else {}
        except Exception as e:
            print(f"❌ Get all users failed: {e}")
            return {}

    def delete_user(self, user_id):
        if not self.initialized:
            return False
        try:
            self.db_ref.child('users').child(self._make_key(user_id)).delete()
            return True
        except Exception as e:
            print(f"❌ Delete user failed: {e}")
            return False

    def save_attendance(self, date, user_id, name, time_str, subject):
        if not self.initialized:
            return False
        try:
            self.db_ref.child('attendance').child(date).child(str(user_id)).set({
                'user_id': user_id,
                'name': name,
                'time': time_str,
                'subject': subject,
                'timestamp': datetime.now().isoformat()
            })
            return True
        except Exception as e:
            print(f"❌ Save attendance failed: {e}")
            return False

    def get_attendance(self, date):
        if not self.initialized:
            return {}
        try:
            attendance = self.db_ref.child('attendance').child(date).get()
            return attendance if attendance else {}
        except Exception as e:
            print(f"❌ Get attendance failed: {e}")
            return {}


class FaceAttendanceSystem:
    def __init__(self):
        self.base_dir = os.path.dirname(os.path.abspath(__file__))
        self.attendance_dir = os.path.join(self.base_dir, "attendance_records")
        self.unknown_dir = os.path.join(self.base_dir, "unknown_faces")

        os.makedirs(self.attendance_dir, exist_ok=True)
        os.makedirs(self.unknown_dir, exist_ok=True)

        self.firebase = FirebaseManager()
        if not self.firebase.initialize():
            print("⚠️  Running OFFLINE mode")

        # Load MediaPipe
        print("Loading MediaPipe face detector...")
        try:
            import mediapipe as mp
            from mediapipe.tasks import python
            from mediapipe.tasks.python import vision

            model_path = os.path.join(self.base_dir, 'detector.tflite')
            if not os.path.exists(model_path):
                print("⬇️  Downloading model...")
                import urllib.request
                urllib.request.urlretrieve(
                    'https://storage.googleapis.com/mediapipe-models/face_detector/blaze_face_short_range/float16/1/blaze_face_short_range.tflite',
                    model_path
                )
                print("✅ Model downloaded!")

            base_options = python.BaseOptions(model_asset_path=model_path)
            options = vision.FaceDetectorOptions(base_options=base_options, min_detection_confidence=0.85)
            self.face_detector = vision.FaceDetector.create_from_options(options)
            self.mp_image = mp.Image
            self.mp_image_format = mp.ImageFormat
            print(f"✓ MediaPipe {mp.__version__} loaded!")
            self.using_mediapipe = True
        except Exception as e:
            print(f"MediaPipe failed: {e} -> using Haarcascade")
            self.face_detector = None
            self.using_mediapipe = False

        self.face_cascade = cv2.CascadeClassifier(
            cv2.data.haarcascades + 'haarcascade_frontalface_default.xml'
        )
        self.recognizer = cv2.face.LBPHFaceRecognizer_create()
        self.load_users()

    def load_users(self):
        print("📥 Loading users from Firebase...")
        users_data = self.firebase.get_all_users()

        if not users_data:
            print("ℹ️  No users in Firebase")
            self.names = {}
            return

        if isinstance(users_data, list):
            print("⚠️  Firebase returned list - converting (data lama)...")
            users_dict = {str(i): v for i, v in enumerate(users_data) if v is not None}
        else:
            users_dict = users_data

        self.names = {}
        faces_list = []
        labels_list = []

        for user_key, user_data in users_dict.items():
            if user_data is None:
                continue
            try:
                user_id_int = int(user_data.get('id', user_key.replace('user_', '')))
                self.names[user_id_int] = user_data['name']

                face_img = self.firebase._decode_face(user_data.get('face_encoding'))
                if face_img is not None:
                    if face_img.shape != (200, 200):
                        face_img = cv2.resize(face_img, (200, 200))
                    faces_list.append(face_img)
                    labels_list.append(user_id_int)
            except Exception as e:
                print(f"⚠️  Skip {user_key}: {e}")
                continue

        if faces_list:
            self.recognizer.train(faces_list, np.array(labels_list))
            print(f"✅ Loaded {len(self.names)} users")
        else:
            print("ℹ️  No valid face encodings")

    def detect_faces_mediapipe(self, frame):
        faces_list = []
        frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        mp_img = self.mp_image(image_format=self.mp_image_format.SRGB, data=frame_rgb)
        result = self.face_detector.detect(mp_img)
        if result.detections:
            for det in result.detections:
                bbox = det.bounding_box
                x = max(0, int(bbox.origin_x))
                y = max(0, int(bbox.origin_y))
                w = min(int(bbox.width), frame.shape[1] - x)
                h = min(int(bbox.height), frame.shape[0] - y)
                score = det.categories[0].score if det.categories else 0.85
                if score >= 0.85:
                    faces_list.append((x, y, w, h))
        return faces_list

    def detect_faces_haarcascade(self, gray):
        return self.face_cascade.detectMultiScale(
            gray, scaleFactor=1.1, minNeighbors=4, minSize=(80, 80)
        )

    def register_face(self, name, person_id):
        """
        DARI SIMPLE VERSION:
        - Haarcascade untuk detection (lebih reliable buat register)
        - SPACE = manual capture
        - A = toggle auto capture
        """
        cap = cv2.VideoCapture(0)
        if not cap.isOpened():
            messagebox.showerror("Error", "Camera tidak bisa dibuka!")
            return

        cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
        cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)

        faces_collected = []
        count = 0
        max_samples = 30
        auto_capture_mode = False
        frame_skip = 0

        print(f"\n📸 Registering: {name} (ID: {person_id})")
        print("SPACE = Manual capture | A = Auto capture | ESC = Batal")

        while count < max_samples:
            ret, frame = cap.read()
            if not ret:
                break

            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            # Pakai Haarcascade untuk register (lebih stabil, dari simple version)
            faces = self.face_cascade.detectMultiScale(
                gray, scaleFactor=1.1, minNeighbors=4, minSize=(80, 80)
            )

            for (x, y, w, h) in faces:
                cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 255, 0), 3)
                cv2.putText(frame, "Wajah Terdeteksi!", (x, y-10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)

            mode_text = "AUTO MODE" if auto_capture_mode else "MANUAL MODE"
            cv2.putText(frame, mode_text, (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 255), 2)
            cv2.putText(frame, f"Foto: {count}/{max_samples}", (10, 60), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
            cv2.putText(frame, f"Nama: {name}", (10, 90), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
            cv2.putText(frame, "SPACE=Capture | A=Auto | ESC=Batal", (10, 120), cv2.FONT_HERSHEY_SIMPLEX, 0.55, (200, 200, 200), 1)

            if len(faces) == 0:
                cv2.putText(frame, "WAJAH TIDAK TERDETEKSI!", (10, frame.shape[0]-20),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)
            else:
                cv2.putText(frame, "Siap! Tekan SPACE atau aktifkan Auto", (10, frame.shape[0]-20),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)

            cv2.imshow(f'Register: {name} - ESC to cancel', frame)
            key = cv2.waitKey(1) & 0xFF

            if key == ord('a') or key == ord('A'):
                auto_capture_mode = not auto_capture_mode
                print(f"{'✓ AUTO MODE ON' if auto_capture_mode else '✗ AUTO MODE OFF'}")

            if key == 32 and len(faces) > 0:  # SPACE
                x, y, w, h = faces[0]
                face_roi = gray[y:y+h, x:x+w]
                faces_collected.append(face_roi)
                count += 1
                print(f"✓ Foto {count}/{max_samples} (MANUAL)")
                cv2.waitKey(100)

            elif auto_capture_mode and len(faces) > 0:
                frame_skip += 1
                if frame_skip >= 5:
                    x, y, w, h = faces[0]
                    face_roi = gray[y:y+h, x:x+w]
                    faces_collected.append(face_roi)
                    count += 1
                    print(f"✓ Foto {count}/{max_samples} (AUTO)")
                    frame_skip = 0
                    cv2.waitKey(50)

            elif key == 27:  # ESC
                cap.release()
                cv2.destroyAllWindows()
                return

        cap.release()
        cv2.destroyAllWindows()

        if len(faces_collected) < 10:
            messagebox.showwarning("Warning", f"Foto terlalu sedikit ({len(faces_collected)}/10 minimum)!\nCoba lagi dengan pencahayaan lebih baik.")
            return

        print("🔄 Training recognizer...")
        # Resize semua ke 200x200 dulu
        faces_resized = [cv2.resize(f, (200, 200)) for f in faces_collected]
        labels = [person_id] * len(faces_resized)

        if self.names:
            self.recognizer.update(faces_resized, np.array(labels))
        else:
            self.recognizer.train(faces_resized, np.array(labels))

        print("☁️  Saving to Firebase...")
        if self.firebase.save_user(person_id, name, faces_resized[0]):
            self.names[person_id] = name
            messagebox.showinfo("Berhasil",
                f"✅ {name} berhasil didaftarkan!\n"
                f"ID: {person_id}\n"
                f"Total foto: {len(faces_collected)}\n"
                f"☁️  Tersimpan di Firebase!")
            print(f"✅ {name} saved to Firebase!")
        else:
            messagebox.showerror("Error", "Gagal menyimpan ke Firebase!")

    def take_attendance(self, subject):
        """
        DARI SIMPLE VERSION:
        - ESC to stop (no time limit)
        - Unknown auto-save with dedup by proximity
        - Counter unknown saved di layar
        - List hadir tampil di kamera
        - MediaPipe untuk detection (lebih akurat buat attendance)
        """
        print(f"\n📋 Attendance: {subject}")

        if not self.names:
            messagebox.showwarning("Warning", "Belum ada wajah terdaftar! Register dulu.")
            return

        cap = cv2.VideoCapture(0)
        if not cap.isOpened():
            messagebox.showerror("Error", "Camera tidak bisa dibuka!")
            return

        attendance_today = {}
        saved_unknowns = {}
        unknown_photo_count = 0

        print(f"📷 Attendance started. Tekan ESC untuk selesai.")
        print(f"Unknown faces akan auto-save ke: {self.unknown_dir}")

        while True:
            ret, frame = cap.read()
            if not ret:
                break

            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

            # Detect dengan MediaPipe kalau ada, fallback haarcascade
            if self.using_mediapipe and self.face_detector:
                raw_faces = self.detect_faces_mediapipe(frame)
                detection_method = "MediaPipe"
            else:
                raw_faces = self.detect_faces_haarcascade(gray)
                detection_method = "Haarcascade"

            for (x, y, w, h) in raw_faces:
                face_roi = gray[y:y+h, x:x+w]
                if face_roi.size == 0:
                    continue
                face_roi_resized = cv2.resize(face_roi, (200, 200))

                person_id = None
                confidence = 100
                is_unknown = True

                if self.names:
                    try:
                        person_id, confidence = self.recognizer.predict(face_roi_resized)
                        if confidence < 70:
                            is_unknown = False
                    except:
                        pass

                if not is_unknown:
                    # Wajah dikenal
                    name = self.names.get(person_id, "Unknown")
                    color = (0, 255, 0)

                    if person_id not in attendance_today:
                        time_now = datetime.now().strftime("%H:%M:%S")
                        attendance_today[person_id] = {'name': name, 'time': time_now, 'id': person_id}
                        self.firebase.save_attendance(
                            datetime.now().strftime("%Y-%m-%d"), person_id, name, time_now, subject
                        )
                        print(f"✅ {name} hadir! ({time_now})")

                    cv2.rectangle(frame, (x, y), (x+w, y+h), color, 2)
                    cv2.putText(frame, f"{name} ({int(confidence)})",
                                (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2)
                else:
                    # Wajah tidak dikenal - auto save dengan dedup
                    center_x = x + w // 2
                    center_y = y + h // 2

                    is_duplicate = False
                    for saved_data in saved_unknowns.values():
                        if 'center' in saved_data:
                            saved_cx, saved_cy = saved_data['center']
                            distance = np.sqrt((center_x - saved_cx)**2 + (center_y - saved_cy)**2)
                            if distance < 150:
                                is_duplicate = True
                                break

                    temp_id = f"{center_x//100}_{center_y//100}"
                    if temp_id not in saved_unknowns and not is_duplicate:
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        unknown_photo_count += 1
                        photo_filename = f"unknown_{timestamp}_{unknown_photo_count}.jpg"
                        photo_path = os.path.join(self.unknown_dir, photo_filename)

                        padding = 20
                        y1 = max(0, y - padding)
                        y2 = min(frame.shape[0], y + h + padding)
                        x1 = max(0, x - padding)
                        x2 = min(frame.shape[1], x + w + padding)
                        cv2.imwrite(photo_path, frame[y1:y2, x1:x2])

                        saved_unknowns[temp_id] = {
                            'photo_path': photo_path,
                            'center': (center_x, center_y)
                        }
                        print(f"📸 Unknown saved: {photo_filename}")

                    color = (0, 165, 255)
                    cv2.rectangle(frame, (x, y), (x+w, y+h), color, 3)
                    cv2.putText(frame, "UNKNOWN", (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2)
                    cv2.putText(frame, "Register nanti", (x, y-28), cv2.FONT_HERSHEY_SIMPLEX, 0.45, color, 1)

            # HUD info
            cv2.putText(frame, f"Subject: {subject} [{detection_method}]", (10, 30),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
            cv2.putText(frame, f"Hadir: {len(attendance_today)}", (10, 60),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
            cv2.putText(frame, f"Unknown saved: {len(saved_unknowns)}", (10, 90),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 165, 255), 2)
            cv2.putText(frame, "ESC = Selesai", (10, 120),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 0), 2)

            # List hadir di layar
            y_offset = 150
            cv2.putText(frame, "HADIR:", (10, y_offset), cv2.FONT_HERSHEY_SIMPLEX, 0.55, (0, 255, 0), 1)
            for idx, (pid, info) in enumerate(list(attendance_today.items())[:6]):
                y_offset += 22
                cv2.putText(frame, f"- {info['name']}", (10, y_offset),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 1)

            cv2.imshow('Attendance - ESC to finish', frame)

            if cv2.waitKey(1) & 0xFF == 27:  # ESC
                break

        cap.release()
        cv2.destroyAllWindows()
        self.save_attendance_excel(attendance_today, subject)

    def save_attendance_excel(self, attendance_dict, subject):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            messagebox.showwarning("Warning", "Install openpyxl:\npip install openpyxl")
            return

        today = datetime.now().strftime("%Y-%m-%d")
        filename = f"{subject}_{today}_{datetime.now().strftime('%H-%M-%S')}.xlsx"
        filepath = os.path.join(self.attendance_dir, filename)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance"

        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        for col, header in enumerate(['No', 'ID', 'Nama', 'Waktu', 'Subject', 'Tanggal'], 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        for width, col in zip([10, 15, 30, 15, 35, 20], 'ABCDEF'):
            ws.column_dimensions[col].width = width

        for idx, (pid, info) in enumerate(attendance_dict.items(), 1):
            for col, val in enumerate([idx, pid, info['name'], info['time'], subject, today], 1):
                cell = ws.cell(row=idx+1, column=col, value=val)
                cell.border = border
                cell.alignment = Alignment(
                    horizontal='center' if col in [1, 2, 4, 6] else 'left',
                    vertical='center'
                )

        wb.save(filepath)

        summary = (f"📊 Attendance Selesai!\n\nSubject: {subject}\nTanggal: {today}\n"
                   f"Total Hadir: {len(attendance_dict)}\n\nHadir:\n")
        for pid, info in attendance_dict.items():
            summary += f"  • {info['name']} ({info['time']})\n"
        summary += f"\n✅ Tersimpan: {filename}\n☁️  Synced ke Firebase!"

        messagebox.showinfo("Selesai", summary)
        print(summary)

    def delete_registered_person(self):
        if not self.names:
            messagebox.showinfo("Info", "Tidak ada orang yang terdaftar!")
            return

        del_window = tk.Toplevel()
        del_window.title("Delete Registered Person")
        del_window.geometry("600x500")
        del_window.configure(bg="#1e1e1e")

        tk.Label(del_window, text=f"Hapus Orang Terdaftar ({len(self.names)} total)",
                 font=("Arial", 16, "bold"), bg="#1e1e1e", fg="#ff0000").pack(pady=10)

        list_frame = tk.Frame(del_window, bg="#1e1e1e")
        list_frame.pack(pady=10, padx=20, fill="both", expand=True)

        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side="right", fill="y")

        lb = tk.Listbox(list_frame, font=("Arial", 12), bg="#333333", fg="yellow",
                        selectmode=tk.SINGLE, yscrollcommand=scrollbar.set, height=15)
        lb.pack(fill="both", expand=True)
        scrollbar.config(command=lb.yview)

        person_list = []
        for pid, name in sorted(self.names.items()):
            lb.insert(tk.END, f"ID: {pid} - {name}")
            person_list.append(pid)

        def delete_selected():
            sel = lb.curselection()
            if not sel:
                messagebox.showwarning("Warning", "Pilih orang yang mau dihapus!")
                return
            pid = person_list[sel[0]]
            name = self.names[pid]
            if messagebox.askyesno("Konfirmasi",
                                   f"Hapus {name}?\nID: {pid}\n\n⚠️ Akan dihapus dari Firebase juga!"):
                if self.firebase.delete_user(pid):
                    del self.names[pid]
                    messagebox.showinfo("Berhasil", f"✅ {name} berhasil dihapus!")
                    print(f"🗑️  Deleted: {name} (ID: {pid})")
                    del_window.destroy()
                    self.delete_registered_person()
                else:
                    messagebox.showerror("Error", "Gagal hapus dari Firebase!")

        def delete_all():
            if messagebox.askyesno("Konfirmasi",
                                   f"Hapus SEMUA {len(self.names)} orang?\n\n⚠️ Tidak bisa di-undo!"):
                for pid in list(self.names.keys()):
                    self.firebase.delete_user(pid)
                self.names = {}
                messagebox.showinfo("Berhasil", "✅ Semua data dihapus!")
                print("🗑️  All users deleted")
                del_window.destroy()

        btn_frame = tk.Frame(del_window, bg="#1e1e1e")
        btn_frame.pack(pady=20)

        for col, (text, cmd, color) in enumerate([
            ("🗑️ Hapus Dipilih", delete_selected, "#cc0000"),
            ("💣 Hapus Semua",   delete_all,       "#990000"),
            ("❌ Tutup",         del_window.destroy, "#666666"),
        ]):
            tk.Button(btn_frame, text=text, command=cmd, font=("Arial", 12, "bold"),
                      bg=color, fg="white", width=18, height=2).grid(row=0, column=col, padx=10)

    def register_unknown_faces(self):
        """
        DARI SIMPLE VERSION:
        - Name entry langsung di window (bukan dialog popup)
        - Tombol Delete Photo
        - Enter key buat save
        """
        if not os.path.exists(self.unknown_dir):
            messagebox.showinfo("Info", "Tidak ada foto unknown yang tersimpan!")
            return

        photo_files = [f for f in os.listdir(self.unknown_dir) if f.endswith('.jpg')]
        if not photo_files:
            messagebox.showinfo("Info", "Tidak ada foto unknown!")
            return

        reg_window = tk.Toplevel()
        reg_window.title(f"Register Unknown Faces ({len(photo_files)} foto)")
        reg_window.geometry("800x650")
        reg_window.configure(bg="#1e1e1e")

        current_index = [0]
        registered_count = [0]

        tk.Label(reg_window, text=f"Register Unknown Faces ({len(photo_files)} foto)",
                 font=("Arial", 18, "bold"), bg="#1e1e1e", fg="#00ff00").pack(pady=10)

        img_label = tk.Label(reg_window, bg="#1e1e1e")
        img_label.pack(pady=10)

        info_label = tk.Label(reg_window, text="", font=("Arial", 12), bg="#1e1e1e", fg="#cccccc")
        info_label.pack(pady=5)

        # Name entry langsung di window (bukan dialog - dari simple version)
        name_frame = tk.Frame(reg_window, bg="#1e1e1e")
        name_frame.pack(pady=10)

        tk.Label(name_frame, text="Nama:", font=("Arial", 14), bg="#1e1e1e", fg="yellow").pack(side="left", padx=5)

        name_entry = tk.Entry(name_frame, font=("Arial", 14), width=25, bg="#333333", fg="yellow",
                              insertbackground="yellow")
        name_entry.pack(side="left", padx=5)

        # ID entry
        id_frame = tk.Frame(reg_window, bg="#1e1e1e")
        id_frame.pack(pady=5)

        tk.Label(id_frame, text="ID:   ", font=("Arial", 14), bg="#1e1e1e", fg="yellow").pack(side="left", padx=5)

        id_entry = tk.Entry(id_frame, font=("Arial", 14), width=10, bg="#333333", fg="yellow",
                            insertbackground="yellow")
        id_entry.pack(side="left", padx=5)

        tk.Label(id_frame, text="(kosongkan = auto ID)", font=("Arial", 10), bg="#1e1e1e", fg="#888888").pack(side="left", padx=5)

        def load_photo():
            if current_index[0] >= len(photo_files):
                messagebox.showinfo("Selesai",
                    f"✅ Selesai!\n\nDiregistrasi: {registered_count[0]}/{len(photo_files)}")
                reg_window.destroy()
                return

            photo_file = photo_files[current_index[0]]
            photo_path = os.path.join(self.unknown_dir, photo_file)

            img = Image.open(photo_path)
            img = img.resize((300, 300), Image.LANCZOS)
            photo = ImageTk.PhotoImage(img)
            img_label.config(image=photo)
            img_label.image = photo

            info_label.config(text=f"Foto {current_index[0]+1} dari {len(photo_files)} | {photo_file}")
            name_entry.delete(0, 'end')
            id_entry.delete(0, 'end')
            name_entry.focus()

        def save_and_next():
            name = name_entry.get().strip()
            if not name:
                messagebox.showwarning("Warning", "Masukkan nama dulu!")
                return

            # Ambil ID atau auto-generate
            id_text = id_entry.get().strip()
            if id_text:
                try:
                    new_id = int(id_text)
                except ValueError:
                    messagebox.showwarning("Warning", "ID harus angka!")
                    return
            else:
                new_id = 9000 + len(self.names)

            if new_id in self.names:
                messagebox.showwarning("Warning", f"ID {new_id} sudah terdaftar untuk: {self.names[new_id]}")
                return

            photo_file = photo_files[current_index[0]]
            photo_path = os.path.join(self.unknown_dir, photo_file)

            face_img = cv2.imread(photo_path, cv2.IMREAD_GRAYSCALE)
            face_resized = cv2.resize(face_img, (200, 200))

            face_copies = [face_resized] * 15
            labels_copies = [new_id] * 15

            if self.names:
                self.recognizer.update(face_copies, np.array(labels_copies))
            else:
                self.recognizer.train(face_copies, np.array(labels_copies))

            if self.firebase.save_user(new_id, name, face_resized):
                self.names[new_id] = name
                os.remove(photo_path)
                photo_files.pop(current_index[0])
                registered_count[0] += 1
                print(f"✓ {name} (ID: {new_id}) registered & saved to Firebase!")
                messagebox.showinfo("Berhasil", f"✅ {name} berhasil didaftarkan!\nID: {new_id}")
                load_photo()
            else:
                messagebox.showerror("Error", "Gagal menyimpan ke Firebase!")

        def delete_photo():
            """DARI SIMPLE VERSION: hapus foto tanpa register"""
            if messagebox.askyesno("Konfirmasi", "Hapus foto ini?\n(Foto akan dihapus permanent)"):
                photo_file = photo_files[current_index[0]]
                photo_path = os.path.join(self.unknown_dir, photo_file)
                os.remove(photo_path)
                photo_files.pop(current_index[0])
                print(f"🗑️  Deleted: {photo_file}")
                load_photo()

        def skip():
            current_index[0] += 1
            load_photo()

        btn_frame = tk.Frame(reg_window, bg="#1e1e1e")
        btn_frame.pack(pady=15)

        for col, (text, cmd, color, width) in enumerate([
            ("✅ Save & Next",   save_and_next, "#00aa00", 14),
            ("🗑️ Delete Photo",  delete_photo,  "#cc0000", 14),
            ("⏭️ Skip",          skip,          "#cc6600", 10),
            ("❌ Tutup",         reg_window.destroy, "#666666", 10),
        ]):
            tk.Button(btn_frame, text=text, command=cmd, font=("Arial", 13, "bold"),
                      bg=color, fg="white", width=width, height=2).grid(row=0, column=col, padx=8)

        # Enter key buat save (dari simple version)
        name_entry.bind('<Return>', lambda e: save_and_next())
        id_entry.bind('<Return>', lambda e: save_and_next())

        load_photo()


class AttendanceGUI:
    def __init__(self):
        self.system = FaceAttendanceSystem()
        self.window = tk.Tk()
        self.window.title("Face Attendance System - Firebase")
        self.window.geometry("500x600")
        self.window.configure(bg="#1e1e1e")
        self.create_widgets()

    def create_widgets(self):
        for w in self.window.winfo_children():
            w.destroy()

        tk.Label(self.window, text="📸 FACE ATTENDANCE",
                 font=("Arial", 24, "bold"), bg="#1e1e1e", fg="#00ff00").pack(pady=20)

        status = "🟢 Online" if self.system.firebase.initialized else "🔴 Offline"
        tk.Label(self.window, text=f"☁️ Firebase Edition | {status}",
                 font=("Arial", 11), bg="#1e1e1e", fg="#cccccc").pack(pady=3)

        btn_frame = tk.Frame(self.window, bg="#1e1e1e")
        btn_frame.pack(pady=20)

        buttons = [
            ("➕ Register New Face",     self.register_new_face,    "#0066cc"),
            ("📋 Take Attendance",       self.take_attendance,      "#00aa00"),
            ("📂 Buka Folder Records",   self.open_records,         "#cc6600"),
            ("👤 Register Unknown",      self.register_unknowns,    "#9900cc"),
            ("📸 Buka Foto Unknown",     self.open_unknown_folder,  "#006699"),
            ("🗑️ Hapus Semua Unknown",   self.delete_all_unknown,   "#990000"),
            ("👤🗑️ Hapus Terdaftar",     self.delete_registered,    "#cc0066"),
            ("❌ Keluar",               self.window.quit,           "#cc0000"),
        ]

        for idx, (text, cmd, color) in enumerate(buttons):
            tk.Button(btn_frame, text=text, command=cmd, font=("Arial", 13, "bold"),
                      bg=color, fg="white", width=24, height=2,
                      cursor="hand2").grid(row=idx, column=0, pady=5, padx=10)

        tk.Label(self.window, text=f"Terdaftar: {len(self.system.names)} orang",
                 font=("Arial", 10), bg="#1e1e1e", fg="#888888").pack(side="bottom", pady=10)

    def register_new_face(self):
        name = simpledialog.askstring("Register", "Masukkan Nama:")
        if not name:
            return
        pid = simpledialog.askinteger("Register", "Masukkan ID (angka):")
        if not pid:
            return
        if pid in self.system.names:
            messagebox.showwarning("Warning", f"ID {pid} sudah terdaftar: {self.system.names[pid]}")
            return
        self.system.register_face(name, pid)
        self.create_widgets()

    def take_attendance(self):
        subject = simpledialog.askstring("Attendance", "Masukkan Nama Subject/Acara:")
        if subject:
            self.system.take_attendance(subject)

    def open_records(self):
        import subprocess
        subprocess.Popen(f'explorer "{self.system.attendance_dir}"')

    def open_unknown_folder(self):
        import subprocess
        if os.path.exists(self.system.unknown_dir):
            subprocess.Popen(f'explorer "{self.system.unknown_dir}"')
        else:
            messagebox.showinfo("Info", "Belum ada folder unknown_faces!")

    def delete_all_unknown(self):
        photos = [f for f in os.listdir(self.system.unknown_dir) if f.endswith('.jpg')] \
                 if os.path.exists(self.system.unknown_dir) else []
        if not photos:
            messagebox.showinfo("Info", "Tidak ada foto unknown!")
            return
        if messagebox.askyesno("Konfirmasi", f"Hapus SEMUA {len(photos)} foto unknown?\n\n⚠️ Tidak bisa di-undo!"):
            for p in photos:
                os.remove(os.path.join(self.system.unknown_dir, p))
            messagebox.showinfo("Berhasil", f"✅ {len(photos)} foto dihapus!")

    def delete_registered(self):
        self.system.delete_registered_person()

    def register_unknowns(self):
        self.system.register_unknown_faces()

    def run(self):
        self.window.mainloop()


if __name__ == "__main__":
    print("=" * 60)
    print("FACE ATTENDANCE SYSTEM - FIREBASE + SIMPLE FEATURES MERGED")
    print("=" * 60)
    AttendanceGUI().run()