"""
SIMPLE FACE ATTENDANCE SYSTEM
Sistem absensi wajah yang gampang dipake
"""

import cv2
import numpy as np
import os
import pickle
import csv
from datetime import datetime
import tkinter as tk
from tkinter import messagebox, simpledialog
from PIL import Image, ImageTk

class FaceAttendanceSystem:
    def __init__(self):
        self.base_dir = os.path.dirname(os.path.abspath(__file__))
        self.data_dir = os.path.join(self.base_dir, "face_data")
        self.attendance_dir = os.path.join(self.base_dir, "attendance_records")
        
        # Create folders
        os.makedirs(self.data_dir, exist_ok=True)
        os.makedirs(self.attendance_dir, exist_ok=True)
        
        # Files
        self.face_data_file = os.path.join(self.data_dir, "faces.pkl")
        self.names_file = os.path.join(self.data_dir, "names.pkl")
        
        # Load MediaPipe face detector
        print("Loading MediaPipe face detector...")
        try:
            import mediapipe as mp
            self.mp_face_detection = mp.solutions.face_detection
            self.face_detector = self.mp_face_detection.FaceDetection(
                model_selection=1,  # 1 = full range
                min_detection_confidence=0.85  # 85% - VERY STRICT!
            )
            print(f"✓ MediaPipe {mp.__version__} loaded! (Confidence: 85%)")
        except Exception as e:
            print(f"MediaPipe load failed: {e}")
            print("Falling back to Haarcascade...")
            self.face_detector = None
        
        # Haarcascade as backup
        self.face_cascade = cv2.CascadeClassifier(
            cv2.data.haarcascades + 'haarcascade_frontalface_default.xml'
        )
        
        # Load face recognizer
        self.recognizer = cv2.face.LBPHFaceRecognizer_create()
        
        # Load existing data
        self.load_data()
        
    def delete_registered_person(self):
        """Delete a registered person from database"""
        if not self.names:
            messagebox.showinfo("Info", "Tidak ada orang yang terdaftar!")
            return
        
        # Create deletion window
        del_window = tk.Toplevel()
        del_window.title("Delete Registered Person")
        del_window.geometry("600x500")
        del_window.configure(bg="#1e1e1e")
        
        # Title
        title_label = tk.Label(
            del_window,
            text=f"Delete Registered Person ({len(self.names)} total)",
            font=("Arial", 16, "bold"),
            bg="#1e1e1e",
            fg="#ff0000"
        )
        title_label.pack(pady=10)
        
        # List frame with scrollbar
        list_frame = tk.Frame(del_window, bg="#1e1e1e")
        list_frame.pack(pady=10, padx=20, fill="both", expand=True)
        
        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side="right", fill="y")
        
        # Listbox
        person_listbox = tk.Listbox(
            list_frame,
            font=("Arial", 12),
            bg="#333333",
            fg="yellow",
            selectmode=tk.SINGLE,
            yscrollcommand=scrollbar.set,
            height=15
        )
        person_listbox.pack(fill="both", expand=True)
        scrollbar.config(command=person_listbox.yview)
        
        # Populate listbox
        person_list = []
        for person_id, name in sorted(self.names.items()):
            display_text = f"ID: {person_id} - {name}"
            person_listbox.insert(tk.END, display_text)
            person_list.append(person_id)
        
        def delete_selected():
            """Delete selected person"""
            selection = person_listbox.curselection()
            if not selection:
                messagebox.showwarning("Warning", "Pilih orang yang mau dihapus!")
                return
            
            idx = selection[0]
            person_id = person_list[idx]
            person_name = self.names[person_id]
            
            if messagebox.askyesno(
                "Confirm Delete",
                f"Hapus {person_name}?\n\nID: {person_id}\n\n⚠️ Data face training akan dihapus permanent!"
            ):
                # Remove from names dict
                del self.names[person_id]
                
                # Rebuild face data without this person
                if os.path.exists(self.face_data_file):
                    with open(self.face_data_file, 'rb') as f:
                        face_data = pickle.load(f)
                    
                    # Filter out faces with this ID
                    new_faces = []
                    new_labels = []
                    for i, label in enumerate(face_data['labels']):
                        if label != person_id:
                            new_faces.append(face_data['faces'][i])
                            new_labels.append(label)
                    
                    if new_faces:
                        # Retrain with remaining faces
                        self.recognizer.train(new_faces, np.array(new_labels))
                        self.save_data(new_faces, new_labels)
                    else:
                        # No faces left, delete files
                        os.remove(self.face_data_file)
                        if os.path.exists(self.names_file):
                            os.remove(self.names_file)
                else:
                    # Just save names
                    with open(self.names_file, 'wb') as f:
                        pickle.dump(self.names, f)
                
                messagebox.showinfo("Success", f"✅ {person_name} berhasil dihapus!")
                print(f"🗑️  Deleted: {person_name} (ID: {person_id})")
                
                # Refresh window
                del_window.destroy()
                self.delete_registered_person()
        
        def delete_all():
            """Delete all registered persons"""
            if messagebox.askyesno(
                "Confirm Delete All",
                f"Hapus SEMUA registered persons?\n\nTotal: {len(self.names)} orang\n\n⚠️ Tidak bisa di-undo!"
            ):
                # Delete all files
                if os.path.exists(self.face_data_file):
                    os.remove(self.face_data_file)
                if os.path.exists(self.names_file):
                    os.remove(self.names_file)
                
                self.names = {}
                
                messagebox.showinfo("Success", f"✅ Semua data berhasil dihapus!")
                print(f"🗑️  All registered persons deleted")
                
                del_window.destroy()
        
        # Buttons
        btn_frame = tk.Frame(del_window, bg="#1e1e1e")
        btn_frame.pack(pady=20)
        
        btn_delete = tk.Button(
            btn_frame,
            text="🗑️ Delete Selected",
            command=delete_selected,
            font=("Arial", 12, "bold"),
            bg="#cc0000",
            fg="white",
            width=18,
            height=2
        )
        btn_delete.grid(row=0, column=0, padx=10)
        
        btn_delete_all = tk.Button(
            btn_frame,
            text="💣 Delete All",
            command=delete_all,
            font=("Arial", 12, "bold"),
            bg="#990000",
            fg="white",
            width=18,
            height=2
        )
        btn_delete_all.grid(row=0, column=1, padx=10)
        
        btn_close = tk.Button(
            btn_frame,
            text="❌ Close",
            command=del_window.destroy,
            font=("Arial", 12, "bold"),
            bg="#666666",
            fg="white",
            width=18,
            height=2
        )
        btn_close.grid(row=0, column=2, padx=10)
    
    def register_unknown_faces(self):
        """Register unknown faces from saved photos"""
        unknown_faces_dir = os.path.join(self.base_dir, "unknown_faces")
        
        if not os.path.exists(unknown_faces_dir):
            messagebox.showinfo("Info", "Tidak ada foto unknown yang tersimpan!")
            return
        
        # Get all unknown photos
        photo_files = [f for f in os.listdir(unknown_faces_dir) if f.endswith('.jpg')]
        
        if not photo_files:
            messagebox.showinfo("Info", "Tidak ada foto unknown yang tersimpan!")
            return
        
        # Create registration window
        reg_window = tk.Toplevel()
        reg_window.title("Register Unknown Faces")
        reg_window.geometry("800x600")
        reg_window.configure(bg="#1e1e1e")
        
        current_index = [0]  # Use list to make it mutable in nested function
        registered_count = [0]
        
        # Title
        title_label = tk.Label(
            reg_window,
            text=f"Register Unknown Faces ({len(photo_files)} photos)",
            font=("Arial", 18, "bold"),
            bg="#1e1e1e",
            fg="#00ff00"
        )
        title_label.pack(pady=10)
        
        # Image display
        img_label = tk.Label(reg_window, bg="#1e1e1e")
        img_label.pack(pady=20)
        
        # Info label
        info_label = tk.Label(
            reg_window,
            text="",
            font=("Arial", 12),
            bg="#1e1e1e",
            fg="#cccccc"
        )
        info_label.pack(pady=5)
        
        # Name entry
        name_frame = tk.Frame(reg_window, bg="#1e1e1e")
        name_frame.pack(pady=10)
        
        tk.Label(
            name_frame,
            text="Nama:",
            font=("Arial", 14),
            bg="#1e1e1e",
            fg="yellow"
        ).pack(side="left", padx=5)
        
        name_entry = tk.Entry(
            name_frame,
            font=("Arial", 14),
            width=30,
            bg="#333333",
            fg="yellow"
        )
        name_entry.pack(side="left", padx=5)
        
        def load_photo():
            if current_index[0] >= len(photo_files):
                messagebox.showinfo(
                    "Done", 
                    f"✅ Selesai!\n\nRegistered: {registered_count[0]}/{len(photo_files)}"
                )
                reg_window.destroy()
                return
            
            photo_file = photo_files[current_index[0]]
            photo_path = os.path.join(unknown_faces_dir, photo_file)
            
            # Load and display image
            img = Image.open(photo_path)
            img = img.resize((300, 300), Image.LANCZOS)
            photo = ImageTk.PhotoImage(img)
            img_label.config(image=photo)
            img_label.image = photo
            
            # Update info
            info_label.config(
                text=f"Photo {current_index[0]+1} of {len(photo_files)}\nFile: {photo_file}"
            )
            
            # Clear name entry
            name_entry.delete(0, 'end')
            name_entry.focus()
        
        def save_and_next():
            name = name_entry.get().strip()
            
            if not name:
                messagebox.showwarning("Warning", "Masukkan nama dulu!")
                return
            
            # Get current photo
            photo_file = photo_files[current_index[0]]
            photo_path = os.path.join(unknown_faces_dir, photo_file)
            
            # Load face image
            face_img = cv2.imread(photo_path)
            gray = cv2.cvtColor(face_img, cv2.COLOR_BGR2GRAY)
            
            # Generate new ID
            new_id = 9000 + len(self.names)
            
            # Register this person
            self.names[new_id] = name
            
            # Train with this face (add multiple copies)
            face_copies = [gray] * 15
            
            # Load existing faces
            if os.path.exists(self.face_data_file):
                with open(self.face_data_file, 'rb') as f:
                    existing_data = pickle.load(f)
                all_faces = existing_data['faces'] + face_copies
                all_labels = existing_data['labels'] + [new_id] * 15
            else:
                all_faces = face_copies
                all_labels = [new_id] * 15
            
            # Retrain
            self.recognizer.train(all_faces, np.array(all_labels))
            self.save_data(all_faces, all_labels)
            
            print(f"✓ {name} registered! (ID: {new_id})")
            
            # Delete photo
            os.remove(photo_path)
            
            registered_count[0] += 1
            current_index[0] += 1
            
            load_photo()
        
        def skip():
            current_index[0] += 1
            load_photo()
        
        def delete_photo():
            """Delete current photo without registering"""
            if messagebox.askyesno("Confirm Delete", "Hapus foto ini?\n(Foto akan dihapus permanent)"):
                # Get current photo
                photo_file = photo_files[current_index[0]]
                photo_path = os.path.join(unknown_faces_dir, photo_file)
                
                # Delete photo
                os.remove(photo_path)
                print(f"🗑️  Deleted: {photo_file}")
                
                current_index[0] += 1
                load_photo()
        
        # Buttons
        btn_frame = tk.Frame(reg_window, bg="#1e1e1e")
        btn_frame.pack(pady=20)
        
        btn_save = tk.Button(
            btn_frame,
            text="✅ Save & Next",
            command=save_and_next,
            font=("Arial", 14, "bold"),
            bg="#00aa00",
            fg="white",
            width=15,
            height=2
        )
        btn_save.grid(row=0, column=0, padx=10)
        
        btn_delete = tk.Button(
            btn_frame,
            text="🗑️ Delete Photo",
            command=delete_photo,
            font=("Arial", 14, "bold"),
            bg="#cc0000",
            fg="white",
            width=15,
            height=2
        )
        btn_delete.grid(row=0, column=1, padx=10)
        
        btn_skip = tk.Button(
            btn_frame,
            text="⏭️ Skip",
            command=skip,
            font=("Arial", 14, "bold"),
            bg="#cc6600",
            fg="white",
            width=15,
            height=2
        )
        btn_skip.grid(row=0, column=2, padx=10)
        
        btn_close = tk.Button(
            btn_frame,
            text="❌ Close",
            command=reg_window.destroy,
            font=("Arial", 14, "bold"),
            bg="#666666",
            fg="white",
            width=15,
            height=2
        )
        btn_close.grid(row=0, column=3, padx=10)
        
        # Bind Enter key to save
        name_entry.bind('<Return>', lambda e: save_and_next())
        
        # Load first photo
        load_photo()
    
    def load_data(self):
        """Load trained face data"""
        if os.path.exists(self.face_data_file) and os.path.exists(self.names_file):
            with open(self.face_data_file, 'rb') as f:
                face_data = pickle.load(f)
            with open(self.names_file, 'rb') as f:
                self.names = pickle.load(f)
            
            # Train recognizer with loaded data
            if len(face_data['faces']) > 0:
                self.recognizer.train(face_data['faces'], np.array(face_data['labels']))
                print(f"✅ Loaded {len(self.names)} registered faces")
        else:
            self.names = {}
            print("ℹ️  No existing data found")
    
    def save_data(self, faces, labels):
        """Save face data"""
        face_data = {'faces': faces, 'labels': labels}
        with open(self.face_data_file, 'wb') as f:
            pickle.dump(face_data, f)
        with open(self.names_file, 'wb') as f:
            pickle.dump(self.names, f)
    
    def register_face(self, name, person_id):
        """Register new face"""
        cap = cv2.VideoCapture(0)
        if not cap.isOpened():
            messagebox.showerror("Error", "Camera tidak bisa dibuka!")
            return False
        
        # Set camera properties for better detection
        cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
        cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
        
        faces_collected = []
        count = 0
        max_samples = 30  # 30 foto cukup
        auto_capture_mode = False
        frame_skip = 0
        
        print(f"\n📸 Mengambil foto untuk: {name} (ID: {person_id})")
        print("SPACE = Manual capture | A = Auto capture | ESC = Batal")
        
        while count < max_samples:
            ret, frame = cap.read()
            if not ret:
                break
            
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            # More sensitive detection parameters
            faces = self.face_cascade.detectMultiScale(
                gray, 
                scaleFactor=1.1,  # Lebih sensitif
                minNeighbors=4,    # Lebih sensitif
                minSize=(80, 80)   # Minimum face size
            )
            
            # Draw face rectangles
            for (x, y, w, h) in faces:
                cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 255, 0), 3)
                cv2.putText(frame, f"Wajah Terdeteksi!", (x, y-10),
                           cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)
            
            # Display instructions
            mode_text = "AUTO MODE" if auto_capture_mode else "MANUAL MODE"
            cv2.putText(frame, mode_text, (10, 30),
                       cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 255), 2)
            cv2.putText(frame, f"Foto: {count}/{max_samples}", (10, 60),
                       cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
            cv2.putText(frame, "SPACE=Capture | A=Auto | ESC=Batal", (10, 90),
                       cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)
            
            # Status message
            if len(faces) == 0:
                cv2.putText(frame, "WAJAH TIDAK TERDETEKSI!", (10, 450),
                           cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)
            else:
                cv2.putText(frame, "Siap! Tekan SPACE atau aktifkan Auto", (10, 450),
                           cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
            
            cv2.imshow('Register Face - ' + name, frame)
            
            key = cv2.waitKey(1) & 0xFF
            
            # Toggle auto-capture mode
            if key == ord('a') or key == ord('A'):
                auto_capture_mode = not auto_capture_mode
                print(f"{'✓ AUTO MODE AKTIF' if auto_capture_mode else '✗ AUTO MODE OFF'}")
            
            # Manual capture with SPACE
            if key == 32 and len(faces) > 0:
                x, y, w, h = faces[0]
                face_roi = gray[y:y+h, x:x+w]
                faces_collected.append(face_roi)
                count += 1
                print(f"✓ Foto {count}/{max_samples} captured (MANUAL)")
                cv2.waitKey(100)  # Small delay to show feedback
            
            # Auto-capture mode
            elif auto_capture_mode and len(faces) > 0:
                frame_skip += 1
                # Capture every 5 frames (slower capture)
                if frame_skip >= 5:
                    x, y, w, h = faces[0]
                    face_roi = gray[y:y+h, x:x+w]
                    faces_collected.append(face_roi)
                    count += 1
                    print(f"✓ Foto {count}/{max_samples} captured (AUTO)")
                    frame_skip = 0
                    cv2.waitKey(100)
            
            # ESC to cancel
            elif key == 27:
                cap.release()
                cv2.destroyAllWindows()
                return False
        
        cap.release()
        cv2.destroyAllWindows()
        
        if len(faces_collected) < 10:
            messagebox.showwarning("Warning", "Foto terlalu sedikit! Minimal 10 foto.")
            return False
        
        # Load existing data
        if os.path.exists(self.face_data_file):
            with open(self.face_data_file, 'rb') as f:
                existing_data = pickle.load(f)
            all_faces = existing_data['faces'] + faces_collected
            all_labels = existing_data['labels'] + [person_id] * len(faces_collected)
        else:
            all_faces = faces_collected
            all_labels = [person_id] * len(faces_collected)
        
        # Save name mapping
        self.names[person_id] = name
        
        # Train recognizer
        self.recognizer.train(all_faces, np.array(all_labels))
        
        # Save data
        self.save_data(all_faces, all_labels)
        
        messagebox.showinfo("Success", f"✅ {name} berhasil didaftarkan!\nTotal foto: {len(faces_collected)}")
        return True
    
    def take_attendance(self, subject="General"):
        """Take attendance - unknown faces are auto-saved for later registration"""
        cap = cv2.VideoCapture(0)
        if not cap.isOpened():
            messagebox.showerror("Error", "Camera tidak bisa dibuka!")
            return
        
        # Create attendance file - LANGSUNG EXCEL!
        today = datetime.now().strftime("%Y-%m-%d")
        time_now = datetime.now().strftime("%H-%M-%S")
        filename = f"{subject}_{today}_{time_now}.xlsx"
        filepath = os.path.join(self.attendance_dir, filename)
        
        # Folder untuk unknown faces
        unknown_faces_dir = os.path.join(self.base_dir, "unknown_faces")
        os.makedirs(unknown_faces_dir, exist_ok=True)
        
        attendance_list = {}
        saved_unknowns = {}  # Track saved unknown photos
        unknown_photo_count = 0
        
        print(f"\n📋 Mengambil absensi: {subject}")
        print("Tekan ESC untuk selesai")
        print(f"Unknown faces akan auto-save ke: {unknown_faces_dir}")
        
        # Import openpyxl untuk Excel
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            print("Installing openpyxl...")
            import subprocess
            subprocess.check_call(['pip', 'install', 'openpyxl', '--break-system-packages'])
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance"
        
        # Header style
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        headers = ['No', 'ID', 'Nama', 'Waktu', 'Subject', 'Tanggal']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Set column widths
        ws.column_dimensions['A'].width = 10  # No
        ws.column_dimensions['B'].width = 15  # ID
        ws.column_dimensions['C'].width = 30  # Nama
        ws.column_dimensions['D'].width = 15  # Waktu
        ws.column_dimensions['E'].width = 35  # Subject
        ws.column_dimensions['F'].width = 20  # Tanggal
        
        while True:
            ret, frame = cap.read()
            if not ret:
                break
            
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces_list = []
            
            # Use MediaPipe if available, otherwise Haarcascade
            if self.face_detector:
                # MediaPipe detection
                frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                results = self.face_detector.process(frame_rgb)
                
                if results.detections:
                    for detection in results.detections:
                        bboxC = detection.location_data.relative_bounding_box
                        ih, iw, _ = frame.shape
                        
                        x = max(0, int(bboxC.xmin * iw))
                        y = max(0, int(bboxC.ymin * ih))
                        w = min(int(bboxC.width * iw), iw - x)
                        h = min(int(bboxC.height * ih), ih - y)
                        
                        confidence_score = detection.score[0]
                        
                        # MediaPipe AI quality filter - VERY STRICT (85%)
                        if confidence_score >= 0.85:
                            faces_list.append({
                                'box': (x, y, w, h),
                                'confidence': confidence_score,
                                'method': 'MediaPipe'
                            })
            else:
                # Haarcascade fallback
                faces = self.face_cascade.detectMultiScale(
                    gray, 
                    scaleFactor=1.2,
                    minNeighbors=6,
                    minSize=(100, 100),
                    flags=cv2.CASCADE_SCALE_IMAGE
                )
                
                for (x, y, w, h) in faces:
                    # Variance check
                    face_roi_check = gray[y:y+h, x:x+w]
                    variance = np.var(face_roi_check)
                    if variance < 150:
                        continue
                    
                    # Aspect ratio check
                    aspect_ratio = w / h
                    if aspect_ratio < 0.7 or aspect_ratio > 1.3:
                        continue
                    
                    faces_list.append({
                        'box': (x, y, w, h),
                        'confidence': 0.85,
                        'method': 'Haarcascade'
                    })
            
            # Process each detected face
            for face_data in faces_list:
                x, y, w, h = face_data['box']
                confidence_score = face_data['confidence']
                detection_method = face_data['method']
                
                face_roi = gray[y:y+h, x:x+w]
                
                person_id = None
                confidence = 100
                is_unknown = True
                
                # Try to recognize if we have trained data
                if self.names and face_roi.size > 0:
                    try:
                        person_id, confidence = self.recognizer.predict(face_roi)
                        if confidence < 70:
                            is_unknown = False
                    except:
                        pass
                
                if not is_unknown:
                    # Known person - AUTO DETECT
                    name = self.names.get(person_id, "Unknown")
                    color = (0, 255, 0)
                    text = f"{name} ({int(confidence)})"
                    
                    # Record attendance (once per person)
                    if person_id not in attendance_list:
                        attendance_list[person_id] = {
                            'name': name,
                            'time': datetime.now().strftime("%H:%M:%S")
                        }
                        
                        # Write to Excel immediately
                        row_number = len(attendance_list) + 1
                        current_time = datetime.now().strftime("%H:%M:%S")
                        
                        ws.cell(row=row_number, column=1, value=len(attendance_list))
                        ws.cell(row=row_number, column=2, value=person_id)
                        ws.cell(row=row_number, column=3, value=name)
                        ws.cell(row=row_number, column=4, value=current_time)
                        ws.cell(row=row_number, column=5, value=subject)
                        ws.cell(row=row_number, column=6, value=today)
                        
                        # Apply styling
                        for col_num in range(1, 7):
                            cell = ws.cell(row=row_number, column=col_num)
                            cell.border = border
                            if col_num in [1, 2, 4, 6]:
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                            else:
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                        
                        wb.save(filepath)
                        print(f"✓ {name} hadir! (Auto-Detect)")
                    
                    cv2.rectangle(frame, (x, y), (x+w, y+h), color, 2)
                    cv2.putText(frame, text, (x, y-10),
                               cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2)
                else:
                    # Unknown person - AUTO SAVE FOTO!
                    # Better duplicate detection using center position
                    
                    center_x = x + w // 2
                    center_y = y + h // 2
                    temp_id = f"{center_x//100}_{center_y//100}"  # Larger grid to prevent duplicates
                    
                    # Check if this face is too close to already saved unknowns
                    is_duplicate = False
                    for saved_id, saved_data in saved_unknowns.items():
                        # Calculate distance from saved positions
                        if 'center' in saved_data:
                            saved_cx, saved_cy = saved_data['center']
                            distance = np.sqrt((center_x - saved_cx)**2 + (center_y - saved_cy)**2)
                            
                            # If within 150 pixels, consider it the same person
                            if distance < 150:
                                is_duplicate = True
                                break
                    
                    if temp_id not in saved_unknowns and not is_duplicate:
                        # All checks passed - save photo
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        unknown_photo_count += 1
                        photo_filename = f"unknown_{timestamp}_{unknown_photo_count}.jpg"
                        photo_path = os.path.join(unknown_faces_dir, photo_filename)
                        
                        # Save face image with some padding
                        padding = 20
                        y1 = max(0, y - padding)
                        y2 = min(frame.shape[0], y + h + padding)
                        x1 = max(0, x - padding)
                        x2 = min(frame.shape[1], x + w + padding)
                        
                        face_img = frame[y1:y2, x1:x2]
                        cv2.imwrite(photo_path, face_img)
                        
                        saved_unknowns[temp_id] = {
                            'photo_path': photo_path,
                            'timestamp': timestamp,
                            'center': (center_x, center_y)  # Store center for duplicate check
                        }
                        
                        print(f"📸 Unknown face saved: {photo_filename} ({detection_method} - confidence: {confidence_score:.2f})")
                    
                    # Show orange box
                    color = (0, 165, 255)  # Orange
                    text = f"UNKNOWN ({int(confidence_score*100)}%)"
                    
                    cv2.rectangle(frame, (x, y), (x+w, y+h), color, 3)
                    cv2.putText(frame, text, (x, y-10),
                               cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2)
                    cv2.putText(frame, "Register nanti", (x, y-30),
                               cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 2)
            
            # Display info
            detection_mode = "MediaPipe" if self.face_detector else "Haarcascade"
            cv2.putText(frame, f"Subject: {subject} [{detection_mode}]", (10, 30),
                       cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
            cv2.putText(frame, f"Faces detected: {len(faces_list)}", (10, 60),
                       cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 255), 2)
            cv2.putText(frame, f"Present: {len(attendance_list)}", (10, 90),
                       cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
            cv2.putText(frame, f"Unknown saved: {len(saved_unknowns)}", (10, 120),
                       cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 165, 255), 2)
            cv2.putText(frame, "ESC = Selesai", (10, 150),
                       cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 0), 2)
            
            # Show list of present people on screen
            y_offset = 180
            cv2.putText(frame, "HADIR:", (10, y_offset),
                       cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)
            for idx, (pid, info) in enumerate(attendance_list.items()):
                if idx < 5:  # Show max 5 names on screen
                    y_offset += 25
                    cv2.putText(frame, f"- {info['name']}", (10, y_offset),
                               cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 1)
            
            cv2.imshow('Taking Attendance', frame)
            
            if cv2.waitKey(1) & 0xFF == 27:  # ESC
                break
        
        cap.release()
        cv2.destroyAllWindows()
        
        # Final save
        wb.save(filepath)
        
        # Show summary
        summary = f"📊 Attendance Summary\n\n"
        summary += f"Subject: {subject}\n"
        summary += f"Date: {today}\n"
        summary += f"Total Present: {len(attendance_list)}\n\n"
        summary += "Present:\n"
        for pid, info in attendance_list.items():
            summary += f"  • {info['name']} ({info['time']})\n"
        summary += f"\n✅ Saved to Excel: {filename}"
        
        messagebox.showinfo("Attendance Complete", summary)
        print(summary)


class AttendanceGUI:
    def __init__(self):
        self.system = FaceAttendanceSystem()
        self.window = tk.Tk()
        self.window.title("Simple Face Attendance System")
        self.window.geometry("500x400")
        self.window.configure(bg="#1e1e1e")
        
        self.create_widgets()
    
    def create_widgets(self):
        # Title
        title = tk.Label(
            self.window,
            text="📸 FACE ATTENDANCE",
            font=("Arial", 24, "bold"),
            bg="#1e1e1e",
            fg="#00ff00"
        )
        title.pack(pady=20)
        
        subtitle = tk.Label(
            self.window,
            text="Simple & Easy Attendance System",
            font=("Arial", 12),
            bg="#1e1e1e",
            fg="#cccccc"
        )
        subtitle.pack(pady=5)
        
        # Button frame
        btn_frame = tk.Frame(self.window, bg="#1e1e1e")
        btn_frame.pack(pady=40)
        
        # Register button
        btn_register = tk.Button(
            btn_frame,
            text="➕ Register New Face",
            command=self.register_new_face,
            font=("Arial", 14, "bold"),
            bg="#0066cc",
            fg="white",
            width=20,
            height=2,
            cursor="hand2"
        )
        btn_register.grid(row=0, column=0, pady=10, padx=10)
        
        # Attendance button
        btn_attendance = tk.Button(
            btn_frame,
            text="📋 Take Attendance",
            command=self.take_attendance,
            font=("Arial", 14, "bold"),
            bg="#00aa00",
            fg="white",
            width=20,
            height=2,
            cursor="hand2"
        )
        btn_attendance.grid(row=1, column=0, pady=10, padx=10)
        
        # View records button
        btn_view = tk.Button(
            btn_frame,
            text="📂 Open Records Folder",
            command=self.open_records,
            font=("Arial", 14, "bold"),
            bg="#cc6600",
            fg="white",
            width=20,
            height=2,
            cursor="hand2"
        )
        btn_view.grid(row=2, column=0, pady=10, padx=10)
        
        # Register unknown button
        btn_register_unknown = tk.Button(
            btn_frame,
            text="👤 Register Unknown Faces",
            command=self.register_unknowns,
            font=("Arial", 14, "bold"),
            bg="#9900cc",
            fg="white",
            width=20,
            height=2,
            cursor="hand2"
        )
        btn_register_unknown.grid(row=3, column=0, pady=10, padx=10)
        
        # Open unknown folder button
        btn_open_unknown = tk.Button(
            btn_frame,
            text="📸 Open Unknown Photos",
            command=self.open_unknown_folder,
            font=("Arial", 14, "bold"),
            bg="#006699",
            fg="white",
            width=20,
            height=2,
            cursor="hand2"
        )
        btn_open_unknown.grid(row=4, column=0, pady=10, padx=10)
        
        # Delete all unknown button
        btn_delete_all = tk.Button(
            btn_frame,
            text="🗑️ Delete All Unknown",
            command=self.delete_all_unknown,
            font=("Arial", 14, "bold"),
            bg="#990000",
            fg="white",
            width=20,
            height=2,
            cursor="hand2"
        )
        btn_delete_all.grid(row=5, column=0, pady=10, padx=10)
        
        # Delete registered person button (NEW!)
        btn_delete_registered = tk.Button(
            btn_frame,
            text="👤🗑️ Delete Registered",
            command=self.delete_registered,
            font=("Arial", 14, "bold"),
            bg="#cc0066",
            fg="white",
            width=20,
            height=2,
            cursor="hand2"
        )
        btn_delete_registered.grid(row=6, column=0, pady=10, padx=10)
        
        # Exit button
        btn_exit = tk.Button(
            btn_frame,
            text="❌ Exit",
            command=self.window.quit,
            font=("Arial", 14, "bold"),
            bg="#cc0000",
            fg="white",
            width=20,
            height=2,
            cursor="hand2"
        )
        btn_exit.grid(row=7, column=0, pady=10, padx=10)
        
        # Info label
        info = tk.Label(
            self.window,
            text=f"Registered: {len(self.system.names)} people",
            font=("Arial", 10),
            bg="#1e1e1e",
            fg="#888888"
        )
        info.pack(side="bottom", pady=10)
    
    def register_new_face(self):
        # Get name
        name = simpledialog.askstring("Register", "Masukkan Nama:")
        if not name:
            return
        
        # Get ID
        person_id = simpledialog.askinteger("Register", "Masukkan ID (angka):")
        if not person_id:
            return
        
        # Check if ID exists
        if person_id in self.system.names:
            messagebox.showwarning("Warning", f"ID {person_id} sudah terdaftar untuk: {self.system.names[person_id]}")
            return
        
        # Register
        self.system.register_face(name, person_id)
        
        # Update info
        self.window.destroy()
        self.__init__()
    
    def take_attendance(self):
        subject = simpledialog.askstring("Attendance", "Masukkan Nama Subject/Acara:")
        if subject:
            self.system.take_attendance(subject)
    
    def open_records(self):
        import subprocess
        subprocess.Popen(f'explorer "{self.system.attendance_dir}"')
    
    def open_unknown_folder(self):
        """Open unknown faces folder"""
        import subprocess
        unknown_dir = os.path.join(self.system.base_dir, "unknown_faces")
        if os.path.exists(unknown_dir):
            subprocess.Popen(f'explorer "{unknown_dir}"')
        else:
            messagebox.showinfo("Info", "Folder unknown_faces belum ada!\nBelum ada foto unknown yang tersimpan.")
    
    def delete_all_unknown(self):
        """Delete all unknown photos"""
        unknown_dir = os.path.join(self.system.base_dir, "unknown_faces")
        
        if not os.path.exists(unknown_dir):
            messagebox.showinfo("Info", "Tidak ada foto unknown!")
            return
        
        photo_files = [f for f in os.listdir(unknown_dir) if f.endswith('.jpg')]
        
        if not photo_files:
            messagebox.showinfo("Info", "Tidak ada foto unknown!")
            return
        
        if messagebox.askyesno(
            "Confirm Delete All",
            f"Hapus SEMUA foto unknown?\n\nTotal: {len(photo_files)} foto\n\n⚠️ Tidak bisa di-undo!"
        ):
            for photo_file in photo_files:
                photo_path = os.path.join(unknown_dir, photo_file)
                os.remove(photo_path)
            
            messagebox.showinfo("Success", f"✅ {len(photo_files)} foto berhasil dihapus!")
            print(f"🗑️  Deleted {len(photo_files)} unknown photos")
    
    def delete_registered(self):
        """Open delete registered person window"""
        self.system.delete_registered_person()
    
    def register_unknowns(self):
        """Open unknown face registration window"""
        self.system.register_unknown_faces()
    
    def run(self):
        self.window.mainloop()


if __name__ == "__main__":
    print("=" * 50)
    print("SIMPLE FACE ATTENDANCE SYSTEM")
    print("=" * 50)
    app = AttendanceGUI()
    app.run()