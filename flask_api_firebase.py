"""
FLASK REST API - FACE ATTENDANCE SYSTEM WITH FIREBASE
API untuk connect dengan frontend (React, Vue, etc)
Firebase Cloud Database Edition
"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import cv2
import numpy as np
import os
import json
import base64
from datetime import datetime
import firebase_admin
from firebase_admin import credentials, db

app = Flask(__name__)
CORS(app)

# Initialize paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ATTENDANCE_DIR = os.path.join(BASE_DIR, "attendance_records")
UNKNOWN_DIR = os.path.join(BASE_DIR, "unknown_faces")

os.makedirs(ATTENDANCE_DIR, exist_ok=True)
os.makedirs(UNKNOWN_DIR, exist_ok=True)

# Initialize Firebase
firebase_initialized = False
db_ref = None

try:
    config_path = os.path.join(BASE_DIR, 'firebase-config.json')
    
    if os.path.exists(config_path):
        with open(config_path, 'r') as f:
            config = json.load(f)
        
        if config.get('type') != 'REPLACE_WITH_YOUR_SERVICE_ACCOUNT_KEY':
            if not firebase_admin._apps:
                cred = credentials.Certificate(config_path)
                firebase_admin.initialize_app(cred, {
                    'databaseURL': config.get('databaseURL')
                })
            
            db_ref = db.reference()
            firebase_initialized = True
            print("✅ Firebase connected!")
        else:
            print("⚠️  Firebase config is template. Please configure it.")
    else:
        print("⚠️  firebase-config.json not found!")
        
except Exception as e:
    print(f"❌ Firebase initialization failed: {e}")

# Load MediaPipe
try:
    import mediapipe as mp
    from mediapipe.tasks import python
    from mediapipe.tasks.python import vision
    
    model_path = os.path.join(BASE_DIR, 'detector.tflite')
    if not os.path.exists(model_path):
        print("⬇️  Downloading face detection model...")
        import urllib.request
        model_url = 'https://storage.googleapis.com/mediapipe-models/face_detector/blaze_face_short_range/float16/1/blaze_face_short_range.tflite'
        urllib.request.urlretrieve(model_url, model_path)
        print("✅ Model downloaded!")
    
    base_options = python.BaseOptions(model_asset_path=model_path)
    options = vision.FaceDetectorOptions(
        base_options=base_options,
        min_detection_confidence=0.85
    )
    face_detector = vision.FaceDetector.create_from_options(options)
    mp_image = mp.Image
    mp_image_format = mp.ImageFormat
    using_mediapipe = True
    print("✅ MediaPipe loaded!")
except Exception as e:
    print(f"⚠️  MediaPipe failed: {e}")
    face_detector = None
    using_mediapipe = False

# Haarcascade fallback
face_cascade = cv2.CascadeClassifier(
    cv2.data.haarcascades + 'haarcascade_frontalface_default.xml'
)

# Face recognizer
recognizer = cv2.face.LBPHFaceRecognizer_create()

# Load users from Firebase
names = {}

def load_users_from_firebase():
    """Load all users from Firebase"""
    global names
    
    if not firebase_initialized:
        return
    
    try:
        users_ref = db_ref.child('users')
        users_data = users_ref.get()
        
        if not users_data:
            print("ℹ️  No users in Firebase")
            return
        
        names = {}
        faces_list = []
        labels_list = []
        
        for user_id, user_data in users_data.items():
            user_id_int = int(user_id)
            names[user_id_int] = user_data['name']
            
            if 'face_encoding' in user_data:
                face_encoding = np.array(user_data['face_encoding'], dtype=np.uint8)
                faces_list.append(face_encoding)
                labels_list.append(user_id_int)
        
        if faces_list:
            recognizer.train(faces_list, np.array(labels_list))
            print(f"✅ Loaded {len(names)} users from Firebase")
            
    except Exception as e:
        print(f"❌ Failed to load users: {e}")

# Load users on startup
load_users_from_firebase()

# Helper functions
def base64_to_image(base64_string):
    """Convert base64 string to OpenCV image"""
    if ',' in base64_string:
        base64_string = base64_string.split(',')[1]
    
    img_data = base64.b64decode(base64_string)
    img_array = np.frombuffer(img_data, np.uint8)
    img = cv2.imdecode(img_array, cv2.IMREAD_COLOR)
    return img

def image_to_base64(img):
    """Convert OpenCV image to base64 string"""
    _, buffer = cv2.imencode('.jpg', img)
    img_base64 = base64.b64encode(buffer).decode('utf-8')
    return f"data:image/jpeg;base64,{img_base64}"

def detect_faces(frame):
    """Detect faces in frame"""
    faces_list = []
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    
    if using_mediapipe and face_detector:
        frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        mp_img = mp_image(image_format=mp_image_format.SRGB, data=frame_rgb)
        detection_result = face_detector.detect(mp_img)
        
        if detection_result.detections:
            for detection in detection_result.detections:
                bbox = detection.bounding_box
                x = int(bbox.origin_x)
                y = int(bbox.origin_y)
                w = int(bbox.width)
                h = int(bbox.height)
                
                x = max(0, x)
                y = max(0, y)
                w = min(w, frame.shape[1] - x)
                h = min(h, frame.shape[0] - y)
                
                confidence_score = detection.categories[0].score if detection.categories else 0.85
                
                if confidence_score >= 0.85:
                    faces_list.append({
                        'box': (x, y, w, h),
                        'confidence': confidence_score
                    })
    else:
        faces = face_cascade.detectMultiScale(
            gray, scaleFactor=1.2, minNeighbors=6, minSize=(100, 100)
        )
        for (x, y, w, h) in faces:
            faces_list.append({
                'box': (x, y, w, h),
                'confidence': 0.85
            })
    
    return faces_list, gray

# ============================================
# API ENDPOINTS
# ============================================

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'firebase': firebase_initialized,
        'mediapipe': using_mediapipe,
        'registered_faces': len(names)
    })

@app.route('/api/detect-face', methods=['POST'])
def detect_face():
    """Detect faces in uploaded image"""
    try:
        data = request.json
        image_data = data.get('image')
        
        if not image_data:
            return jsonify({'error': 'No image provided'}), 400
        
        frame = base64_to_image(image_data)
        faces_list, gray = detect_faces(frame)
        
        for face_data in faces_list:
            x, y, w, h = face_data['box']
            cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 255, 0), 2)
            cv2.putText(frame, "Face detected", (x, y-10),
                       cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)
        
        result_image = image_to_base64(frame)
        
        return jsonify({
            'success': True,
            'faces_detected': len(faces_list),
            'image': result_image
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/recognize', methods=['POST'])
def recognize_faces():
    """Detect AND recognize faces - REAL-TIME"""
    try:
        data = request.json
        image_data = data.get('image')
        confidence_threshold = data.get('confidence_threshold', 70)
        
        if not image_data:
            return jsonify({'error': 'No image provided'}), 400
        
        frame = base64_to_image(image_data)
        faces_list, gray = detect_faces(frame)
        
        result_faces = []
        for face_data in faces_list:
            x, y, w, h = face_data['box']
            face_roi = gray[y:y+h, x:x+w]
            face_roi_resized = cv2.resize(face_roi, (200, 200))
            
            person_id = None
            person_name = "Unknown"
            confidence = 100
            
            if len(names) > 0:
                try:
                    person_id, confidence = recognizer.predict(face_roi_resized)
                    
                    if confidence < confidence_threshold and person_id in names:
                        person_name = names[person_id]
                        color = (0, 255, 0)  # Green
                    else:
                        person_id = None
                        person_name = "Unknown"
                        color = (0, 0, 255)  # Red
                except:
                    person_id = None
                    person_name = "Unknown"
                    color = (0, 0, 255)
            else:
                color = (0, 0, 255)
            
            cv2.rectangle(frame, (x, y), (x+w, y+h), color, 2)
            
            display_text = f"{person_name}"
            if person_id:
                display_text += f" ({int(100-confidence)}%)"
            
            cv2.putText(frame, display_text, (x, y-10),
                       cv2.FONT_HERSHEY_SIMPLEX, 0.8, color, 2)
            
            result_faces.append({
                'person_id': int(person_id) if person_id else None,
                'name': person_name,
                'confidence': float(100 - confidence) if person_id else 0,
                'bbox': {'x': x, 'y': y, 'w': w, 'h': h}
            })
        
        result_image = image_to_base64(frame)
        
        return jsonify({
            'success': True,
            'faces': result_faces,
            'image': result_image
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/register', methods=['POST'])
def register_face():
    """Register new face with Firebase"""
    try:
        if not firebase_initialized:
            return jsonify({'error': 'Firebase not configured'}), 500
        
        data = request.json
        name = data.get('name')
        person_id = data.get('person_id')
        image_data = data.get('image')
        
        if not all([name, person_id, image_data]):
            return jsonify({'error': 'Missing required fields'}), 400
        
        if person_id in names:
            return jsonify({'error': f'ID {person_id} already registered'}), 400
        
        frame = base64_to_image(image_data)
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces_list, _ = detect_faces(frame)
        
        if not faces_list:
            return jsonify({'error': 'No face detected in image'}), 400
        
        x, y, w, h = faces_list[0]['box']
        face_roi = gray[y:y+h, x:x+w]
        face_roi_resized = cv2.resize(face_roi, (200, 200))
        
        # Train recognizer
        if names:
            recognizer.update([face_roi_resized], np.array([person_id]))
        else:
            recognizer.train([face_roi_resized], np.array([person_id]))
        
        # Save to Firebase
        face_encoding = face_roi_resized.flatten()
        
        user_ref = db_ref.child('users').child(str(person_id))
        user_ref.set({
            'name': name,
            'id': person_id,
            'face_encoding': face_encoding.tolist(),
            'created_at': datetime.now().isoformat(),
            'updated_at': datetime.now().isoformat()
        })
        
        names[person_id] = name
        
        return jsonify({
            'success': True,
            'message': f'{name} registered successfully',
            'person_id': person_id
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/attendance/take', methods=['POST'])
def take_attendance():
    """Take attendance with Firebase"""
    try:
        if not firebase_initialized:
            return jsonify({'error': 'Firebase not configured'}), 500
        
        data = request.json
        subject = data.get('subject', 'General')
        image_data = data.get('image')
        
        if not image_data:
            return jsonify({'error': 'No image provided'}), 400
        
        frame = base64_to_image(image_data)
        faces_list, gray = detect_faces(frame)
        
        attendance_list = []
        unknown_list = []
        
        date_today = datetime.now().strftime("%Y-%m-%d")
        time_now = datetime.now().strftime("%H:%M:%S")
        
        for face_data in faces_list:
            x, y, w, h = face_data['box']
            face_roi = gray[y:y+h, x:x+w]
            face_roi_resized = cv2.resize(face_roi, (200, 200))
            
            if len(names) > 0:
                person_id, confidence = recognizer.predict(face_roi_resized)
                
                if confidence < 70 and person_id in names:
                    name = names[person_id]
                    
                    # Save to Firebase
                    attendance_ref = db_ref.child('attendance').child(date_today).child(str(person_id))
                    attendance_ref.set({
                        'user_id': person_id,
                        'name': name,
                        'time': time_now,
                        'subject': subject,
                        'timestamp': datetime.now().isoformat()
                    })
                    
                    attendance_list.append({
                        'id': person_id,
                        'name': name,
                        'time': time_now
                    })
                else:
                    unknown_list.append({'time': time_now})
            else:
                unknown_list.append({'time': time_now})
        
        # Save to Excel
        if attendance_list:
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                filename = f"{subject}_{date_today}_{time_now.replace(':', '-')}.xlsx"
                filepath = os.path.join(ATTENDANCE_DIR, filename)
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Attendance"
                
                header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=12)
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                headers = ['No', 'ID', 'Name', 'Time', 'Subject', 'Date']
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                
                ws.column_dimensions['A'].width = 10
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 30
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 35
                ws.column_dimensions['F'].width = 20
                
                for idx, person in enumerate(attendance_list, 1):
                    ws.cell(row=idx+1, column=1, value=idx)
                    ws.cell(row=idx+1, column=2, value=person['id'])
                    ws.cell(row=idx+1, column=3, value=person['name'])
                    ws.cell(row=idx+1, column=4, value=person['time'])
                    ws.cell(row=idx+1, column=5, value=subject)
                    ws.cell(row=idx+1, column=6, value=date_today)
                    
                    for col_num in range(1, 7):
                        cell = ws.cell(row=idx+1, column=col_num)
                        cell.border = border
                
                wb.save(filepath)
            except:
                pass
        
        return jsonify({
            'success': True,
            'subject': subject,
            'date': date_today,
            'attendance': attendance_list,
            'unknown': unknown_list
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/registered', methods=['GET'])
def get_registered():
    """Get all registered persons from Firebase"""
    try:
        if not firebase_initialized:
            return jsonify({'error': 'Firebase not configured'}), 500
        
        users_ref = db_ref.child('users')
        users_data = users_ref.get()
        
        if not users_data:
            return jsonify({'success': True, 'count': 0, 'registered': []})
        
        registered = [
            {'id': int(uid), 'name': udata['name']}
            for uid, udata in users_data.items()
        ]
        
        return jsonify({
            'success': True,
            'count': len(registered),
            'registered': registered
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/registered/<int:person_id>', methods=['DELETE'])
def delete_registered(person_id):
    """Delete registered person from Firebase"""
    try:
        if not firebase_initialized:
            return jsonify({'error': 'Firebase not configured'}), 500
        
        if person_id not in names:
            return jsonify({'error': 'Person not found'}), 404
        
        person_name = names[person_id]
        
        # Delete from Firebase
        user_ref = db_ref.child('users').child(str(person_id))
        user_ref.delete()
        
        del names[person_id]
        
        # Reload users to retrain recognizer
        load_users_from_firebase()
        
        return jsonify({
            'success': True,
            'message': f'{person_name} deleted from Firebase'
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/sync', methods=['POST'])
def sync_firebase():
    """Force sync/reload from Firebase"""
    try:
        load_users_from_firebase()
        
        return jsonify({
            'success': True,
            'message': 'Synced with Firebase',
            'registered_count': len(names)
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    print("=" * 60)
    print("FLASK API - FACE ATTENDANCE SYSTEM WITH FIREBASE")
    print("=" * 60)
    print(f"✅ Firebase: {'Connected' if firebase_initialized else 'Not configured'}")
    print(f"✅ MediaPipe: {'Enabled' if using_mediapipe else 'Disabled'}")
    print(f"✅ Registered faces: {len(names)}")
    print("=" * 60)
    print("Available Endpoints:")
    print("  - GET  /api/health")
    print("  - POST /api/detect-face")
    print("  - POST /api/recognize")
    print("  - POST /api/register")
    print("  - POST /api/attendance/take")
    print("  - GET  /api/registered")
    print("  - DELETE /api/registered/<id>")
    print("  - POST /api/sync")
    print("=" * 60)
    print("🚀 Starting server on http://localhost:5000")
    print("=" * 60)
    
    app.run(host='0.0.0.0', port=5000, debug=True)