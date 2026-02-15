"""
FLASK REST API - FACE ATTENDANCE SYSTEM
API untuk connect dengan frontend (React, Vue, etc)
FULL VERSION dengan /api/recognize endpoint
"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import cv2
import numpy as np
import os
import pickle
import base64
from datetime import datetime
import io
from PIL import Image

# Import face attendance system
import sys
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

app = Flask(__name__)
CORS(app)  # Enable CORS untuk frontend

# Initialize paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "face_data")
ATTENDANCE_DIR = os.path.join(BASE_DIR, "attendance_records")
UNKNOWN_DIR = os.path.join(BASE_DIR, "unknown_faces")

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(ATTENDANCE_DIR, exist_ok=True)
os.makedirs(UNKNOWN_DIR, exist_ok=True)

# Files
FACE_DATA_FILE = os.path.join(DATA_DIR, "faces.pkl")
NAMES_FILE = os.path.join(DATA_DIR, "names.pkl")

# Load MediaPipe
try:
    import mediapipe as mp
    from mediapipe.tasks import python
    from mediapipe.tasks.python import vision
    
    model_path = os.path.join(BASE_DIR, 'detector.tflite')
    if not os.path.exists(model_path):
        print("⬇️  Downloading face detection model...")
        import urllib.request
        model_url = 'https://storage.googleapis.com/mediapipe-models/face_detector/blaze_face_short_range/float16/1/blaze_face_short_range.tflite'
        urllib.request.urlretrieve(model_url, model_path)
        print("✅ Model downloaded!")
    
    base_options = python.BaseOptions(model_asset_path=model_path)
    options = vision.FaceDetectorOptions(
        base_options=base_options,
        min_detection_confidence=0.85
    )
    face_detector = vision.FaceDetector.create_from_options(options)
    mp_image = mp.Image
    mp_image_format = mp.ImageFormat
    using_mediapipe = True
    print("✅ MediaPipe loaded!")
except Exception as e:
    print(f"⚠️  MediaPipe failed, using Haarcascade: {e}")
    face_detector = None
    using_mediapipe = False

# Haarcascade fallback
face_cascade = cv2.CascadeClassifier(
    cv2.data.haarcascades + 'haarcascade_frontalface_default.xml'
)

# Face recognizer
recognizer = cv2.face.LBPHFaceRecognizer_create()

# Load existing data
names = {}
if os.path.exists(FACE_DATA_FILE) and os.path.exists(NAMES_FILE):
    with open(FACE_DATA_FILE, 'rb') as f:
        face_data = pickle.load(f)
    with open(NAMES_FILE, 'rb') as f:
        names = pickle.load(f)
    if len(face_data['faces']) > 0:
        recognizer.train(face_data['faces'], np.array(face_data['labels']))
        print(f"✅ Loaded {len(names)} registered faces")

# Helper functions
def base64_to_image(base64_string):
    """Convert base64 string to OpenCV image"""
    # Remove header if exists
    if ',' in base64_string:
        base64_string = base64_string.split(',')[1]
    
    img_data = base64.b64decode(base64_string)
    img_array = np.frombuffer(img_data, np.uint8)
    img = cv2.imdecode(img_array, cv2.IMREAD_COLOR)
    return img

def image_to_base64(img):
    """Convert OpenCV image to base64 string"""
    _, buffer = cv2.imencode('.jpg', img)
    img_base64 = base64.b64encode(buffer).decode('utf-8')
    return f"data:image/jpeg;base64,{img_base64}"

def detect_faces(frame):
    """Detect faces in frame using MediaPipe or Haarcascade"""
    faces_list = []
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    
    if using_mediapipe and face_detector:
        # MediaPipe detection
        frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        mp_img = mp_image(image_format=mp_image_format.SRGB, data=frame_rgb)
        detection_result = face_detector.detect(mp_img)
        
        if detection_result.detections:
            for detection in detection_result.detections:
                bbox = detection.bounding_box
                x = int(bbox.origin_x)
                y = int(bbox.origin_y)
                w = int(bbox.width)
                h = int(bbox.height)
                
                x = max(0, x)
                y = max(0, y)
                w = min(w, frame.shape[1] - x)
                h = min(h, frame.shape[0] - y)
                
                confidence_score = detection.categories[0].score if detection.categories else 0.85
                
                if confidence_score >= 0.85:
                    faces_list.append({
                        'box': (x, y, w, h),
                        'confidence': confidence_score,
                        'method': 'MediaPipe'
                    })
    else:
        # Haarcascade fallback
        faces = face_cascade.detectMultiScale(
            gray, scaleFactor=1.2, minNeighbors=6, minSize=(100, 100)
        )
        for (x, y, w, h) in faces:
            faces_list.append({
                'box': (x, y, w, h),
                'confidence': 0.85,
                'method': 'Haarcascade'
            })
    
    return faces_list, gray

def save_data(faces, labels):
    """Save face data"""
    face_data = {'faces': faces, 'labels': labels}
    with open(FACE_DATA_FILE, 'wb') as f:
        pickle.dump(face_data, f)
    with open(NAMES_FILE, 'wb') as f:
        pickle.dump(names, f)

# ============================================
# API ENDPOINTS
# ============================================

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'mediapipe': using_mediapipe,
        'registered_faces': len(names)
    })

@app.route('/api/detect-face', methods=['POST'])
def detect_face():
    """Detect faces in uploaded image"""
    try:
        data = request.json
        image_data = data.get('image')
        
        if not image_data:
            return jsonify({'error': 'No image provided'}), 400
        
        # Convert base64 to image
        frame = base64_to_image(image_data)
        
        # Detect faces
        faces_list, gray = detect_faces(frame)
        
        # Draw rectangles on frame
        for face_data in faces_list:
            x, y, w, h = face_data['box']
            cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 255, 0), 2)
            cv2.putText(frame, f"Face detected", (x, y-10),
                       cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)
        
        # Convert back to base64
        result_image = image_to_base64(frame)
        
        return jsonify({
            'success': True,
            'faces_detected': len(faces_list),
            'image': result_image
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/recognize', methods=['POST'])
def recognize_faces():
    """
    *** NEW ENDPOINT ***
    Detect AND recognize faces in image - REAL-TIME MODE
    This endpoint combines detection + recognition in one call
    """
    try:
        data = request.json
        image_data = data.get('image')
        confidence_threshold = data.get('confidence_threshold', 70)
        
        if not image_data:
            return jsonify({'error': 'No image provided'}), 400
        
        # Convert base64 to image
        frame = base64_to_image(image_data)
        
        # Detect faces
        faces_list, gray = detect_faces(frame)
        
        # Process each face
        result_faces = []
        for face_data in faces_list:
            x, y, w, h = face_data['box']
            face_roi = gray[y:y+h, x:x+w]
            
            person_id = None
            confidence = 100
            is_unknown = True
            confidence_level = "UNKNOWN"
            
            # Try to recognize if we have trained data
            if names and face_roi.size > 0:
                try:
                    person_id, confidence = recognizer.predict(face_roi)
                    
                    # Check if confidence is good enough (SAME AS BACKEND PYTHON)
                    if confidence < confidence_threshold:
                        is_unknown = False
                        
                        # Determine confidence level (SAME AS BACKEND)
                        if confidence < 50:
                            confidence_level = "PERFECT"
                        elif confidence < 60:
                            confidence_level = "GOOD"
                        else:
                            confidence_level = "MODERATE"
                except Exception as e:
                    print(f"Recognition error: {e}")
                    pass
            
            # Build face data
            face_info = {
                'box': {'x': x, 'y': y, 'w': w, 'h': h},
                'method': face_data['method'],
                'detection_confidence': float(face_data['confidence'])
            }
            
            if not is_unknown:
                # Known person
                face_info['id'] = int(person_id)
                face_info['name'] = names.get(person_id, "Unknown")
                face_info['confidence'] = int(confidence)
                face_info['confidence_level'] = confidence_level
            else:
                # Unknown person
                face_info['id'] = None
                face_info['name'] = "Unknown"
                face_info['confidence'] = int(confidence)
                face_info['confidence_level'] = "UNKNOWN"
            
            result_faces.append(face_info)
        
        return jsonify({
            'success': True,
            'faces': result_faces,
            'total_detected': len(faces_list),
            'recognized': len([f for f in result_faces if f['name'] != 'Unknown']),
            'unknown': len([f for f in result_faces if f['name'] == 'Unknown'])
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/register', methods=['POST'])
def register_face():
    """Register new person with face images"""
    try:
        data = request.json
        name = data.get('name')
        person_id = data.get('id')
        images = data.get('images', [])  # Array of base64 images
        
        if not name or not person_id:
            return jsonify({'error': 'Name and ID required'}), 400
        
        if person_id in names:
            return jsonify({'error': f'ID {person_id} already registered'}), 400
        
        if len(images) < 5:
            return jsonify({'error': 'At least 5 images required'}), 400
        
        # Process images
        faces_collected = []
        for img_data in images:
            frame = base64_to_image(img_data)
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            
            # Detect face
            faces_list, _ = detect_faces(frame)
            if len(faces_list) > 0:
                x, y, w, h = faces_list[0]['box']
                face_roi = gray[y:y+h, x:x+w]
                faces_collected.append(face_roi)
        
        if len(faces_collected) < 5:
            return jsonify({'error': 'Not enough valid face images'}), 400
        
        # Load existing data
        if os.path.exists(FACE_DATA_FILE):
            with open(FACE_DATA_FILE, 'rb') as f:
                existing_data = pickle.load(f)
            all_faces = existing_data['faces'] + faces_collected
            all_labels = existing_data['labels'] + [person_id] * len(faces_collected)
        else:
            all_faces = faces_collected
            all_labels = [person_id] * len(faces_collected)
        
        # Save and train
        names[person_id] = name
        recognizer.train(all_faces, np.array(all_labels))
        save_data(all_faces, all_labels)
        
        return jsonify({
            'success': True,
            'message': f'{name} registered successfully',
            'id': person_id,
            'name': name,
            'samples': len(faces_collected)
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/attendance/take', methods=['POST'])
def take_attendance():
    """Take attendance from image"""
    try:
        data = request.json
        image_data = data.get('image')
        subject = data.get('subject', 'General')
        
        if not image_data:
            return jsonify({'error': 'No image provided'}), 400
        
        # Convert base64 to image
        frame = base64_to_image(image_data)
        
        # Detect faces
        faces_list, gray = detect_faces(frame)
        
        if len(faces_list) == 0:
            return jsonify({'error': 'No faces detected'}), 400
        
        # Process faces
        attendance_list = []
        unknown_list = []
        
        for face_data in faces_list:
            x, y, w, h = face_data['box']
            face_roi = gray[y:y+h, x:x+w]
            
            person_id = None
            confidence = 100
            is_unknown = True
            
            if names and face_roi.size > 0:
                try:
                    person_id, confidence = recognizer.predict(face_roi)
                    if confidence < 70:  # SAME THRESHOLD AS BACKEND
                        is_unknown = False
                except:
                    pass
            
            current_time = datetime.now().strftime("%H:%M:%S")
            
            if not is_unknown:
                # Known person
                name = names.get(person_id, "Unknown")
                
                # Confidence level
                if confidence < 50:
                    conf_level = "PERFECT"
                elif confidence < 60:
                    conf_level = "GOOD"
                else:
                    conf_level = "MODERATE"
                
                attendance_list.append({
                    'id': int(person_id),
                    'name': name,
                    'time': current_time,
                    'confidence': int(confidence),
                    'confidence_level': conf_level,
                    'method': face_data['method']
                })
            else:
                # Unknown person - save photo
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
                filename = f"unknown_{timestamp}.jpg"
                filepath = os.path.join(UNKNOWN_DIR, filename)
                
                # Save face image
                padding = 20
                y1 = max(0, y - padding)
                y2 = min(frame.shape[0], y + h + padding)
                x1 = max(0, x - padding)
                x2 = min(frame.shape[1], x + w + padding)
                face_img = frame[y1:y2, x1:x2]
                cv2.imwrite(filepath, face_img)
                
                unknown_list.append({
                    'filename': filename,
                    'time': current_time,
                    'method': face_data['method']
                })
        
        # Create Excel file
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            import subprocess
            subprocess.check_call(['pip', 'install', 'openpyxl', '--break-system-packages'])
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        today = datetime.now().strftime("%Y-%m-%d")
        time_now = datetime.now().strftime("%H-%M-%S")
        filename = f"{subject}_{today}_{time_now}.xlsx"
        filepath = os.path.join(ATTENDANCE_DIR, filename)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance"
        
        # Header style
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        headers = ['No', 'ID', 'Nama', 'Waktu', 'Subject', 'Tanggal']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Set column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 35
        ws.column_dimensions['F'].width = 20
        
        # Write data
        for idx, person in enumerate(attendance_list, 1):
            ws.cell(row=idx+1, column=1, value=idx)
            ws.cell(row=idx+1, column=2, value=person['id'])
            ws.cell(row=idx+1, column=3, value=person['name'])
            ws.cell(row=idx+1, column=4, value=person['time'])
            ws.cell(row=idx+1, column=5, value=subject)
            ws.cell(row=idx+1, column=6, value=today)
            
            for col_num in range(1, 7):
                cell = ws.cell(row=idx+1, column=col_num)
                cell.border = border
                if col_num in [1, 2, 4, 6]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
        
        wb.save(filepath)
        
        return jsonify({
            'success': True,
            'subject': subject,
            'date': today,
            'attendance': attendance_list,
            'unknown': unknown_list,
            'file': filename
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/registered', methods=['GET'])
def get_registered():
    """Get all registered persons"""
    try:
        registered = [
            {'id': int(pid), 'name': name}
            for pid, name in names.items()
        ]
        return jsonify({
            'success': True,
            'count': len(registered),
            'registered': registered
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/registered/<int:person_id>', methods=['DELETE'])
def delete_registered(person_id):
    """Delete a registered person"""
    try:
        if person_id not in names:
            return jsonify({'error': 'Person not found'}), 404
        
        person_name = names[person_id]
        del names[person_id]
        
        # Rebuild face data
        if os.path.exists(FACE_DATA_FILE):
            with open(FACE_DATA_FILE, 'rb') as f:
                face_data = pickle.load(f)
            
            new_faces = []
            new_labels = []
            for i, label in enumerate(face_data['labels']):
                if label != person_id:
                    new_faces.append(face_data['faces'][i])
                    new_labels.append(label)
            
            if new_faces:
                recognizer.train(new_faces, np.array(new_labels))
                save_data(new_faces, new_labels)
            else:
                os.remove(FACE_DATA_FILE)
                if os.path.exists(NAMES_FILE):
                    os.remove(NAMES_FILE)
        
        return jsonify({
            'success': True,
            'message': f'{person_name} deleted successfully'
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/unknown', methods=['GET'])
def get_unknown():
    """Get all unknown face photos"""
    try:
        if not os.path.exists(UNKNOWN_DIR):
            return jsonify({'success': True, 'unknown': []})
        
        photos = [f for f in os.listdir(UNKNOWN_DIR) if f.endswith('.jpg')]
        unknown_list = []
        
        for photo in photos:
            photo_path = os.path.join(UNKNOWN_DIR, photo)
            # Read and convert to base64
            img = cv2.imread(photo_path)
            img_base64 = image_to_base64(img)
            
            unknown_list.append({
                'filename': photo,
                'image': img_base64
            })
        
        return jsonify({
            'success': True,
            'count': len(unknown_list),
            'unknown': unknown_list
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/unknown/<filename>', methods=['DELETE'])
def delete_unknown(filename):
    """Delete an unknown face photo"""
    try:
        photo_path = os.path.join(UNKNOWN_DIR, filename)
        if not os.path.exists(photo_path):
            return jsonify({'error': 'Photo not found'}), 404
        
        os.remove(photo_path)
        return jsonify({
            'success': True,
            'message': f'{filename} deleted successfully'
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/save-unknown', methods=['POST'])
def save_unknown_face():
    """Save unknown face from frontend (real-time detection)"""
    try:
        data = request.json
        image_data = data.get('image')
        
        if not image_data:
            return jsonify({'error': 'No image provided'}), 400
        
        # Convert base64 to image
        frame = base64_to_image(image_data)
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        filename = f"unknown_{timestamp}.jpg"
        filepath = os.path.join(UNKNOWN_DIR, filename)
        
        # Save image
        cv2.imwrite(filepath, frame)
        
        return jsonify({
            'success': True,
            'filename': filename,
            'message': 'Unknown face saved successfully'
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/files/<filename>', methods=['GET'])
def download_file(filename):
    """Download attendance file"""
    try:
        filepath = os.path.join(ATTENDANCE_DIR, filename)
        if not os.path.exists(filepath):
            return jsonify({'error': 'File not found'}), 404
        
        return send_file(filepath, as_attachment=True)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    print("=" * 50)
    print("FLASK API - FACE ATTENDANCE SYSTEM")
    print("FULL VERSION with /api/recognize endpoint")
    print("=" * 50)
    print(f"✅ MediaPipe: {'Enabled' if using_mediapipe else 'Disabled (using Haarcascade)'}")
    print(f"✅ Registered faces: {len(names)}")
    print("=" * 50)
    print("Available Endpoints:")
    print("  - GET  /api/health")
    print("  - POST /api/detect-face")
    print("  - POST /api/recognize          [NEW!]")
    print("  - POST /api/register")
    print("  - POST /api/attendance/take")
    print("  - GET  /api/registered")
    print("  - DELETE /api/registered/<id>")
    print("  - GET  /api/unknown")
    print("  - DELETE /api/unknown/<filename>")
    print("=" * 50)
    print("🚀 Starting server on http://localhost:5000")
    print("=" * 50)
    
    app.run(host='0.0.0.0', port=5000, debug=True)